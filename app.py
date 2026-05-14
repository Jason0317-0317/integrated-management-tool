import streamlit as st
import pandas as pd
import io
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
from dotenv import load_dotenv

# 載入環境變數
load_dotenv()

# ========================
# 頁面設定
# ========================
st.set_page_config(page_title="春不老統計系統", layout="wide")

# ========================
# Session State 初始化
# ========================
if "role" not in st.session_state:
    st.session_state.role = None
if "feature" not in st.session_state:
    st.session_state.feature = None
if "show_finance_login" not in st.session_state:
    st.session_state.show_finance_login = False
if "show_manager_login" not in st.session_state:
    st.session_state.show_manager_login = False
if "retry_count" not in st.session_state:
    st.session_state.retry_count = 0
if "lockout_time" not in st.session_state:
    st.session_state.lockout_time = None
if "retry_count_mgr" not in st.session_state:
    st.session_state.retry_count_mgr = 0
if "lockout_time_mgr" not in st.session_state:
    st.session_state.lockout_time_mgr = None

# ========================
# 步驟 1: 身份選擇
# ========================
if st.session_state.role is None:
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        st.markdown("<h1 style='text-align: center;'>春不老統計系統</h1>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; color: #999;'>請選擇您的角色登入系統</p>", unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        col_btn1, col_btn2 = st.columns(2)
        
        with col_btn1:
            if st.button("店長登入", key="manager_btn", use_container_width=True):
                st.session_state.show_manager_login = True
                st.session_state.show_finance_login = False
        
        with col_btn2:
            if st.button("小編登入", key="editor_btn", use_container_width=True):
                st.session_state.role = "editor"
                st.rerun()
        col_btn3, col_btn4 = st.columns(2)
        with col_btn3:
            if st.button("美編登入", key="designer_btn", use_container_width=True):
                st.session_state.role = "designer"
                st.rerun()

        with col_btn4:
            if st.button("財務登入", key="finance_btn", use_container_width=True):
                st.session_state.show_finance_login = True
                st.session_state.show_manager_login = False
        
        if "retry_count" not in st.session_state:
            st.session_state.retry_count = 0
        if "lockout_time" not in st.session_state:
            st.session_state.lockout_time = None
        # --- 店長密碼驗證區塊 ---
        if st.session_state.get("show_manager_login", False):
            st.markdown("<br>", unsafe_allow_html=True)
            is_locked_mgr = False
            
            if st.session_state.retry_count_mgr >= 3:
                time_diff = datetime.now() - st.session_state.lockout_time_mgr
                if time_diff < timedelta(minutes=5):
                    remaining_time = 5 - int(time_diff.total_seconds() // 60)
                    st.error(f"密碼錯誤過多，請休息 {remaining_time} 分鐘再嘗試。")
                    is_locked_mgr = True
                    if st.button("返回", key="mgr_lock_back", use_container_width=True):
                        st.session_state.show_manager_login = False
                        st.rerun()
                else:
                    st.session_state.retry_count_mgr = 0
                    st.session_state.lockout_time_mgr = None

            if not is_locked_mgr:
                mgr_password = st.text_input(
                    f"請輸入店長密碼 (剩餘次數: {3 - st.session_state.retry_count_mgr})",
                    type="password", key="mgr_pwd_input"
                )
                col_m1, col_m2 = st.columns(2)
                with col_m1:
                    if st.button("確認", key="mgr_confirm_btn", use_container_width=True):
                        if mgr_password == "0168":  # 在此設定店長密碼
                            st.session_state.role = "manager"
                            st.session_state.show_manager_login = False
                            st.session_state.retry_count_mgr = 0
                            st.success("店長登入成功！")
                            st.rerun()
                        else:
                            st.session_state.retry_count_mgr += 1
                            if st.session_state.retry_count_mgr >= 3:
                                st.session_state.lockout_time_mgr = datetime.now()
                            st.rerun()
                with col_m2:
                    if st.button("取消", key="mgr_cancel_btn", use_container_width=True):
                        st.session_state.show_manager_login = False
                        st.rerun()
        # 財務密碼驗證區塊
        if st.session_state.get("show_finance_login", False):
            st.markdown("<br>", unsafe_allow_html=True)
            is_locked = False
            if st.session_state.retry_count >= 3:
                # 計算距離上次錯誤過了多久
                time_diff = datetime.now() - st.session_state.lockout_time
                if time_diff < timedelta(minutes=5):
                    remaining_time = 5 - int(time_diff.total_seconds() // 60)
                    st.error(f"密碼錯誤次數過多，請休息 {remaining_time} 分鐘後再嘗試。")
                    is_locked = True
                    if st.button("返回", key="lockout_back_btn", use_container_width=True):
                        st.session_state.show_finance_login = False
                        st.rerun()
                else:
                    # 超過 5 分鐘，重置次數
                    st.session_state.retry_count = 0
                    st.session_state.lockout_time = None
            if not is_locked:
                password = st.text_input(
                    f"請輸入財務密碼 (剩餘嘗試次數: {3 - st.session_state.retry_count})",
                    type="password",
                    key="finance_password_input"
                )
                col_confirm, col_cancel = st.columns(2)
                with col_confirm:
                    if st.button("確認", key="finance_confirm_btn", use_container_width=True):
                        if password == "20260512": # 在此設定財務密碼
                            st.session_state.role = "finance"
                            st.session_state.show_finance_login = False
                            st.session_state.retry_count = 0  # 成功後重置
                            st.session_state.lockout_time = None
                            st.success("登入成功！")
                            st.rerun()
                        else:
                            st.session_state.retry_count += 1
                            if st.session_state.retry_count >= 3:
                                st.session_state.lockout_time = datetime.now()
                                st.error("錯誤達 3 次，系統已暫時鎖定 5 分鐘")
                            else:
                                st.error(f"密碼錯誤！剩餘次數: {3 - st.session_state.retry_count}")
                            st.rerun() # 立即更新頁面顯示錯誤次數
                with col_cancel:
                    if st.button("取消", key="finance_cancel_btn", use_container_width=True):
                        st.session_state.show_finance_login = False
                        st.rerun()
# ========================
# 步驟 2: 功能選擇
# ========================
elif st.session_state.feature is None:
    col_back, col_title = st.columns([1, 10])
    with col_back:
        if st.button("返回"):
            st.session_state.role = None
            st.session_state.feature = None
            st.rerun()
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        if st.session_state.role == "manager":
            st.markdown("<h1 style='text-align: center;'>店長功能選擇</h1>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #999;'>請選擇您要使用的功能</p>", unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # 使用三欄位佈局
            col_f1, col_f2, col_f3 = st.columns(3)
            
            with col_f1:
                if st.button("教練堂數統計", key="lesson_report", use_container_width=True):
                    st.session_state.feature = "lesson_report"
                    st.rerun()
            
            with col_f2:
                if st.button("團績個績統計", key="sales_report", use_container_width=True):
                    st.session_state.feature = "sales_report"
                    st.rerun()
            
            with col_f3:
                if st.button("出勤明細統計", key="attendance_report", use_container_width=True):
                    st.session_state.feature = "attendance_report"
                    st.rerun()
        
        elif st.session_state.role == "editor":
            st.markdown("<h1 style='text-align: center;'>小編功能選擇</h1>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #999;'>請選擇您要使用的功能</p>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            
            if st.button("小編獎金統計", key="editor_bonus", use_container_width=True):
                st.session_state.feature = "editor_bonus"
                st.rerun()
        elif st.session_state.role == "designer":
            st.markdown("<h1 style='text-align: center;'>美編功能選擇</h1>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #999;'>請選擇您要使用的功能</p>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            
            st.info("美編功能開發中，敬請期待。")

        elif st.session_state.role == "finance":
            st.markdown("<h1 style='text-align: center;'>財務功能選擇</h1>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #999;'>請選擇您要使用的功能</p>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            
            if st.button("教練薪資統計", key="coach_salary_btn", use_container_width=True):
                st.session_state.feature = "coach_salary"
                st.rerun()

# ========================
# 步驟 3: 顯示功能
# ========================
else:
    col_back, col_title, col_empty = st.columns([1, 10, 1])
    with col_back:
        if st.button("返回"):
            st.session_state.feature = None
            st.rerun()
    
    # ===== 預約報表統計 =====
    if st.session_state.feature == "lesson_report":
        st.title("教練堂數統計")
        st.markdown("請上傳團體課預約報表。")
        
        # 1. 定義老師排序順序與轉換字典
        NAME_CONVERSION = {
            "意潔": "林意潔", "Cammy": "陳怡廷", "Vivi": "陳秀蓉", "怡廷": "陳怡廷", 
            "秀蓉": "陳秀蓉", "佳蓁": "鍾佳蓁 Rita", "宛婷": "黃宛婷", "WanTing": "黃宛婷",
            "小在": "楊子慧(小在)", "Jae": "楊子慧(小在)", "LOUIS": "許力尹 LOUIS", 
            "許力尹": "許力尹 LOUIS", "顥顥": "顥顥", "睿絃": "洪睿絃", "儒蓁": "紀儒蓁",
            "翎瑋": "李翎瑋", "奕伶": "郭奕伶", "品均": "郭品均", "妍語": "邴妍語", 
            "鈞弼": "張鈞弼", "竣升": "蕭竣升", "萃萃": "紀萃文", "萃文": "紀萃文", 
            "函豫": "李函豫", "Hanny": "遊函豫", "子綺": "尤子綺", "Yuli": "尤子綺", 
            "楷翌": "張楷翌", "Eric": "張楷翌", "懿庭": "侯懿庭", "Yvonne Hou": "侯懿庭", 
            "俐池": "謝俐池", "Grace Hsieh": "謝俐池", "姿菁": "黃姿菁", "郁雯": "籃郁雯", 
            "徐漫": "徐漫", "mandy": "徐漫", "漫漫": "徐漫", "筠馨": "鄭筠馨", 
            "舒涵": "高舒涵", "靜瑜": "邱靜瑜"
        }
        
        # 建立排序參考清單
        TEACHER_ORDER = list(dict.fromkeys(NAME_CONVERSION.values()))
        
        # 名稱轉換函數
        def get_formal_name(raw_name):
            name_str = str(raw_name).strip()
            for key, formal_name in NAME_CONVERSION.items():
                if key.lower() in name_str.lower():
                    return formal_name
            return name_str
        
        # 排序權重函數
        def teacher_sort_key(name):
            if name in TEACHER_ORDER:
                return TEACHER_ORDER.index(name)
            return len(TEACHER_ORDER)
        
        # --- 篩選條件區塊 ---
        st.markdown("### 1. 設定篩選條件")
        col_branch, col_date = st.columns(2)
        
        with col_branch:
            selected_branch = st.selectbox(
                "選擇館別", 
                ["全部", "中山館", "高美館", "義昌館", "巨蛋館"],
                key="branch_select"
            )
        
        with col_date:
            today = datetime.today()
            first_day_of_month = today.replace(day=1)
            date_range = st.date_input(
                "選擇日期區間",
                value=(first_day_of_month, today),
                help="請選取開始與結束日期",
                key="date_range"
            )
        
        if len(date_range) != 2:
            st.warning("請在日曆上選擇完整的開始與結束日期。")
            st.stop()
        
        # --- 檔案上傳區塊 ---
        st.markdown("### 2. 上傳報表檔案")
        uploaded_file = st.file_uploader("選擇原始 Excel 檔案", type=["xlsx", "csv"], key="lesson_uploader")
        
        if uploaded_file is not None:
            try:
                def get_clean_df(file):
                    if file.name.endswith(('.xlsx', '.xls')):
                        temp_df = pd.read_excel(file, header=None, nrows=20)
                        file.seek(0)
                        target_row = 0
                        for i, row in temp_df.iterrows():
                            row_str = " ".join([str(x) for x in row.values])
                            if '課程日期' in row_str or '授課老師' in row_str:
                                target_row = i
                                break
                        return pd.read_excel(file, skiprows=target_row)
                    else:
                        encodings = ['utf-8-sig', 'big5', 'cp950', 'gbk']
                        for enc in encodings:
                            try:
                                file.seek(0)
                                df = pd.read_csv(file, encoding=enc)
                                if '課程日期' not in "".join(df.columns.astype(str)):
                                    file.seek(0)
                                    df = pd.read_csv(file, encoding=enc, skiprows=1)
                                return df
                            except:
                                continue
                        return None
                
                df = get_clean_df(uploaded_file)
                
                if df is None or df.empty:
                    st.error("無法辨識檔案格式或找不到『課程日期』。")
                    st.stop()
                
                df.columns = df.columns.astype(str).str.strip()
                
                def find_col(possible_names):
                    for name in possible_names:
                        for col in df.columns:
                            if name in col and "Unnamed" not in col:
                                return col
                    return None
                
                date_col = find_col(['課程日期', '日期'])
                teacher_col = find_col(['授課老師', '老師'])
                course_col = find_col(['課程名稱', '課程'])
                count_col = find_col(['預約總人數', '預約人數', '人數'])
                duration_col = find_col(['課程時數', '分鐘'])
                branch_col = find_col(['館別', '分館'])
                
                if not date_col or not teacher_col:
                    st.error(f"缺少必要欄位。偵測到的欄位有：{list(df.columns)}")
                    st.stop()
                
                # 資料清洗與篩選
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                df = df.dropna(subset=[date_col])
                
                if branch_col and selected_branch != "全部":
                    df = df[df[branch_col].astype(str).str.contains(selected_branch)]
                
                start_date, end_date = date_range
                df = df[(df[date_col].dt.date >= start_date) & (df[date_col].dt.date <= end_date)]
                
                df[count_col] = pd.to_numeric(df[count_col], errors='coerce').fillna(0)
                df_filtered = df[df[count_col] > 0].copy()
                
                # 轉換正式姓名
                df_filtered['正式姓名'] = df_filtered[teacher_col].apply(get_formal_name)
                
                # 統計邏輯
                stats_items = [
                    '團1人', '團2人', '團3人', '團4人', '團5人', '團6人',
                    '1對2(1.5hr)', '1對2', '1對1(1.5hr)', '1對1', '觀課'
                ]
                
                all_formal_teachers = df_filtered['正式姓名'].unique().tolist()
                df_stats = pd.DataFrame(0, index=all_formal_teachers, columns=stats_items)
                
                for _, row in df_filtered.iterrows():
                    teacher = row['正式姓名']
                    course_name = str(row[course_col]).strip()
                    count = int(row[count_col])
                    duration = row[duration_col] if duration_col else 60
                    
                    if '觀課' in course_name:
                        df_stats.at[teacher, '觀課'] += 1
                    elif '一對一' in course_name:
                        if duration >= 90: df_stats.at[teacher, '1對1(1.5hr)'] += 1
                        else: df_stats.at[teacher, '1對1'] += 1
                    elif '一對二' in course_name:
                        if duration >= 90: df_stats.at[teacher, '1對2(1.5hr)'] += 1
                        else: df_stats.at[teacher, '1對2'] += 1
                    else:
                        if 1 <= count <= 6:
                            col_name = f'團{count}人'
                            df_stats.at[teacher, col_name] += 1
                
                df_stats['小計'] = df_stats.sum(axis=1)
                
                # 排序
                df_stats['sort_key'] = df_stats.index.map(teacher_sort_key)
                df_stats = df_stats.sort_values('sort_key').drop(columns=['sort_key'])
                
                # 合計列
                total_row = df_stats.sum().to_frame().T
                total_row.index = ['合計']
                df_final = pd.concat([df_stats, total_row])
                
                # 轉置
                df_transposed = df_final.T
                
                st.markdown("### 3. 統計結果預覽")
                st.dataframe(df_transposed, use_container_width=True)
                
                # 匯出 Excel
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_transposed.to_excel(writer, sheet_name='堂數統計')
                
                st.download_button(
                    label="下載統計報表 (Excel)",
                    data=output.getvalue(),
                    file_name=f"堂數統計_{selected_branch}_{start_date}_{end_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"處理過程中發生錯誤: {e}")

    # ===== 團績個績統計工具 =====
    elif st.session_state.feature == "sales_report":
        st.title("團績個績統計")
        st.markdown("請上傳交易報表。")
        
        uploaded_file = st.file_uploader("選擇原始 Excel 檔案", type=["xlsx"], key="sales_uploader")
        
        if uploaded_file is not None:
            try:
                df_raw = pd.read_excel(uploaded_file, header=1)
                
                # 建立基礎資料
                df_base = pd.DataFrame()
                df_base["合約建立日期"] = df_raw["交易日期"]
                df_base["業績人員"] = df_raw["銷售人員"]
                df_base["會員姓名"] = df_raw["會員姓名"]
                df_base["開課日期"] = np.nan
                df_base["合約類型"] = df_raw["票券/商品種類"]
                df_base["堂數"] = pd.to_numeric(df_raw["數量"], errors='coerce').fillna(0)
                
                total_price = pd.to_numeric(df_raw["總計(元)"], errors='coerce').fillna(0)
                df_base["合約總價"] = total_price
                df_base["金額(未稅)"] = (total_price / 1.05).round(0)
                
                def get_unit_price(contract_type):
                    contract_type = str(contract_type)
                    if '一對一' in contract_type: return 2300
                    elif '一對二' in contract_type: return 1300
                    elif '一對三' in contract_type: return 1050
                    elif '一對四' in contract_type: return 900
                    else: return 0
                
                df_base["購買合約原價"] = df_base["合約類型"].apply(get_unit_price) * df_base["堂數"]
                
                df_base["銷售折數"] = np.where(
                    df_base["購買合約原價"] != 0,
                    (df_base["金額(未稅)"] / df_base["購買合約原價"]).round(2),
                    0
                )
                
                df_base["活動"] = df_raw["Tag"]
                
                cond = [
                    (df_base["銷售折數"] > 0.8),
                    (df_base["銷售折數"] == 0.8),
                    (df_base["銷售折數"] < 0.8) & (df_base["銷售折數"] > 0)
                ]
                
                bonus_rate = np.select(cond, [0.02, 0.015, 0.01], default=0)
                df_base["業績計算"] = np.select(cond, ["2%", "1.5%", "1%"], default="0%")
                df_base["業績獎金"] = (df_base["金額(未稅)"] * bonus_rate).round(0)
                
                df_base["備註"] = df_raw["備註"].astype(str).replace('nan', '')
                
                target_columns = [
                    "合約建立日期", "業績人員", "會員姓名", "開課日期", "合約類型", 
                    "堂數", "合約總價", "金額(未稅)", "銷售折數", "活動", 
                    "業績計算", "業績獎金", "購買合約原價", "備註"
                ]
                df_base = df_base[target_columns]
                
                df_new = df_base[df_base["備註"].str.contains("新購|首購")].copy()
                df_renew = df_base[df_base["備註"].str.contains("續購")].copy()
                df_trial = df_base[df_base["備註"].str.contains("體驗")].copy()
                df_new = df_new.sort_values(by="業績人員").reset_index(drop=True)
                df_renew = df_renew.sort_values(by="業績人員").reset_index(drop=True)
                df_trial = df_trial.sort_values(by="業績人員").reset_index(drop=True)
                
                def add_subtotals(df):
                    numeric_cols = ["堂數", "合約總價", "金額(未稅)", "業績獎金", "購買合約原價"]
                    result = []
                    for person, group in df.groupby("業績人員", sort=False):
                        result.append(group)
                        subtotal = {col: "" for col in df.columns}
                        subtotal["業績人員"] = f"【{person} 小計】"
                        for col in numeric_cols:
                            if col in df.columns:
                                subtotal[col] = group[col].sum()
                        result.append(pd.DataFrame([subtotal]))
                    return pd.concat(result, ignore_index=True) if result else df
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    add_subtotals(df_new).to_excel(writer, sheet_name='個績', index=False)
                    add_subtotals(df_renew).to_excel(writer, sheet_name='團績', index=False)
                    df_trial.to_excel(writer, sheet_name='體驗', index=False)
                st.download_button(
                    label="下載轉換後報表 (Excel)",
                    data=output.getvalue(),
                    file_name=f"業績報表轉換_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"處理過程中發生錯誤: {e}")

    # ===== 出勤明細統計 =====
    elif st.session_state.feature == "attendance_report":
        st.title("出勤明細統計")
        st.markdown("上傳進出場記錄。")
        
        uploaded_file = st.file_uploader("選擇原始 Excel 檔案", type=["xlsx"], key="attendance_uploader")
        
        if uploaded_file is not None:
            try:
                df = pd.read_excel(uploaded_file)
                
                # 資料清洗
                df['進場日期'] = pd.to_datetime(df['進場日期'], errors='coerce').dt.date
                df['出場日期'] = pd.to_datetime(df['出場日期'], errors='coerce').dt.date
                df['總時間(分)'] = pd.to_numeric(df['總時間(分)'], errors='coerce')
                
                # 統計
                summary = (
                    df.groupby('姓名')
                    .agg(
                        出勤天數=('進場日期', 'nunique'),
                        總時數_分=('總時間(分)', 'sum'),
                    )
                    .reset_index()
                    .sort_values('姓名')
                )
                
                def minutes_to_hhmm(minutes):
                    if pd.isna(minutes): return ""
                    h, m = divmod(int(minutes), 60)
                    return f"{h}:{m:02d}"
                
                summary['總工時(時:分)'] = summary['總時數_分'].apply(minutes_to_hhmm)
                
                # 報表生成邏輯
                if st.button("產生完整報表並下載"):
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # Sheet 1: 出勤明細
                        df_sorted = df.sort_values(['姓名', '進場日期', '進場時間']).reset_index(drop=True)
                        df_sorted['總工時(時:分)'] = df_sorted['總時間(分)'].apply(minutes_to_hhmm)
                        drop_cols = ['進場方案/預約名稱', '進場方式', '出場方式']
                        df_sorted = df_sorted.drop(columns=drop_cols, errors='ignore')
                        df_sorted.to_excel(writer, sheet_name='出勤明細', index=False)
                        
                        # Sheet 2: 月份出勤統計
                        summary.to_excel(writer, sheet_name='月份出勤統計', index=False)
                        
                        # Sheet 3: 月曆排班表
                        df2 = df.copy()
                        df2['日期'] = df2['進場日期'].astype(str).str[:10]
                        pivot = df2.pivot_table(
                            index='姓名', columns='日期', values='總時間(分)', aggfunc='sum'
                        ).fillna(0)
                        pivot.to_excel(writer, sheet_name='月曆排班表')
                        
                    st.download_button(
                        label="點我儲存 Excel 檔案",
                        data=output.getvalue(),
                        file_name=f"出勤明細統計_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"處理過程中發生錯誤: {e}")

    # ===== 小編獎金統計 =====
    elif st.session_state.feature == "editor_bonus":
        st.title("小編獎金統計")
        
        def calculate_bonus(deal_dict, classes, loyalty_dict, upgrade_counts, is_ft, brand_count, revenue_tier, si_to_st):
            # 1. 體驗成交獎金
            d_bonus = (deal_dict["當天"] * 80 + deal_dict["48小時"] * 60 + 
                      deal_dict["7天內"] * 50 + deal_dict["超過7天"] * 0)
            
            # 2. 補位獎金
            c_bonus = classes * 30
            
            # 3. 回流獎金
            l_bonus = (loyalty_dict["10堂"] * 100 + loyalty_dict["20堂"] * 200 + 
                      loyalty_dict["30堂"] * 300 + loyalty_dict["40堂"] * 500)
            
            # 4. 結構升級獎金
            u_bonus = (upgrade_counts["1對2變1對3"] * 100 + 
                      upgrade_counts["團課變期班"] * 150 + 
                      upgrade_counts["包班成立"] * 300)
            
            # 5. 品牌知名度獎金
            base_val = 5 if is_ft else 2
            if brand_count == 0:
                b_bonus = -200
                b_note = "推廣人數為 0"
            elif brand_count < base_val:
                b_bonus = -100
                b_note = f"未達門檻 ({base_val}位)"
            elif brand_count == base_val:
                b_bonus = 0
                b_note = "符合基本門檻"
            else:
                extra_units = (brand_count - base_val) // 5
                b_bonus = extra_units * 200
                b_note = f"加發 {extra_units} 組獎金"
            if si_to_st >= 25:
                s_bonus = ((si_to_st - 20) // 5) * 200
            else:
                s_bonus = 0
            
            # 6. 月高手獎勵
            total_v = (sum(deal_dict.values()) + classes + sum(loyalty_dict.values()) + sum(upgrade_counts.values()) + brand_count + si_to_st)
            if total_v >= 50: m_bonus = 5000
            elif total_v >= 30: m_bonus = 2000
            else: m_bonus = 0
            
            # 7. 個人業績獎金
            rev_map = {"不列入計算": 0, "12萬元": 2000, "24萬元": 4000, "30萬元": 6000}
            r_bonus = rev_map.get(revenue_tier, 0)
            
            total = d_bonus + c_bonus + l_bonus + u_bonus + b_bonus + m_bonus + r_bonus + s_bonus
            return total, total_v, m_bonus, l_bonus, d_bonus, u_bonus, b_bonus, b_note, r_bonus, s_bonus

        def generate_matrix_excel(meta_data, total_v, result, deal_dict, classes, loyalty_dict, upgrade_counts, d_bonus, l_bonus, u_bonus, m_bonus, b_bonus, b_note, emp_type, b_count, r_bonus, r_tier):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                workbook = writer.book
                worksheet = workbook.create_sheet('獎金結算', 0)
                bold_font = Font(bold=True, name='微軟正黑體')
                center_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                info = [
                    ["館別", meta_data["館別"]],
                    ["報表日期", meta_data["報表日期"]],
                    ["小編姓名", meta_data["小編姓名"]],
                    ["員工身份", emp_type]
                ]
                for i, (k, v) in enumerate(info, 1):
                    worksheet.cell(row=i, column=1, value=k).font = bold_font
                    worksheet.cell(row=i, column=2, value=v)
                
                header_row = 7
                headers = ["項目", "筆數", "獎金金額", "備註"]
                for i, h in enumerate(headers, 1):
                    worksheet.cell(row=header_row, column=i, value=h).font = bold_font
                
                data_rows = [
                    ["個人業績獎金", r_tier if r_tier else "不列入計算", r_bonus, ""],
                    ["體驗成交", sum(deal_dict.values()), d_bonus, ""],
                    ["補位獎金", classes, classes * 30, ""],
                    ["SI 轉 ST", si_to_st, s_bonus, "超過20筆起計"],
                    ["回流獎金", sum(loyalty_dict.values()), l_bonus, ""],
                    ["結構升級獎金", sum(upgrade_counts.values()), u_bonus, ""],
                    ["品牌知名度獎金", b_count, b_bonus, b_note],
                    ["月高手獎勵", total_v, m_bonus, f"總轉換筆數: {total_v}"],
                    ["總計", "", result, ""]
                ]
                for r_idx, row_data in enumerate(data_rows, header_row + 1):
                    for c_idx, value in enumerate(row_data, 1):
                        cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                        cell.alignment = center_align
                        cell.border = thin_border
                        if r_idx == header_row + len(data_rows): cell.font = bold_font
                
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 15
                worksheet.column_dimensions['C'].width = 15
                worksheet.column_dimensions['D'].width = 20
            return output.getvalue()
        
        st.markdown("### 基本資訊設定")
        col1, col2, col3, col4 = st.columns(4)
        with col1: gym = st.selectbox("館別", ["義昌館", "高美館", "中山館", "巨蛋館"], key="gym_select")
        with col2: name = st.text_input("小編姓名", "", placeholder="請輸入姓名", key="name_input")
        with col3:
            today = datetime.today()
            first_day = today.replace(day=1)

            date_range = st.date_input(
            "報表日期區間",
            value=(first_day, today),
            key="date_range_input"
        )

        if len(date_range) == 2:
            start_date, end_date = date_range
            date_str = f"{start_date} ~ {end_date}"
        else:
            date_str = ""
        with col4: is_ft = st.selectbox("員工身份", ["正職", "兼職"], key="employment_select") == "正職"
        
        st.divider()
        st.markdown("### 1. 體驗與品牌推廣")
        revenue_tier = st.selectbox("個人業績獎金級別", ["不列入計算", "12萬元", "24萬元", "30萬元"], key="revenue_tier_select")
        st.divider()
        col_a, col_b = st.columns(2)
        with col_a:
            d_today = st.number_input("當天成交(筆)", min_value=0, value=0, key="deal_today")
            d_48h = st.number_input("48小時(筆)", min_value=0, value=0, key="deal_48h")
            d_7d = st.number_input("7天內(筆)", min_value=0, value=0, key="deal_7d")
            d_over7 = st.number_input("超過7天(筆)", min_value=0, value=0, key="deal_over7")
            deal_dict = {"當天": d_today, "48小時": d_48h, "7天內": d_7d, "超過7天": d_over7}
        with col_b:
            brand_input = st.number_input("品牌推廣人數", min_value=0, value=0, key="brand_input")
            extra_cls = st.number_input("補開課程次數", min_value=0, value=0, key="extra_cls")
            si_to_st_input = st.number_input("SI 轉 ST 筆數", min_value=0, value=0, key="si_to_st")
        st.markdown("### 2. 回流與升級項目")
        col_c, col_d = st.columns(2)
        with col_c:
            st.write("回流人數 (STP-T)")
            l_10 = st.number_input("10堂人數", min_value=0, value=0, key="loyalty_10")
            l_20 = st.number_input("20堂人數", min_value=0, value=0, key="loyalty_20")
            l_30 = st.number_input("30堂人數", min_value=0, value=0, key="loyalty_30")
            l_40 = st.number_input("40堂人數", min_value=0, value=0, key="loyalty_40")
            loyalty_dict = {"10堂": l_10, "20堂": l_20, "30堂": l_30, "40堂": l_40}
        with col_d:
            st.write("結構升級次數")
            u_12_13 = st.number_input("1對2變1對3(次)", min_value=0, value=0, key="upgrade_1213")
            u_group = st.number_input("團課變期班(次)", min_value=0, value=0, key="upgrade_group")
            u_class = st.number_input("包班成立(次)", min_value=0, value=0, key="upgrade_class")
            upgrade_dict = {"1對2變1對3": u_12_13, "團課變期班": u_group, "包班成立": u_class}
        
        res = calculate_bonus(deal_dict, extra_cls, loyalty_dict, upgrade_dict, is_ft, brand_input, revenue_tier, si_to_st_input)
        st.divider()
        main_col1, main_col2 = st.columns(2)
        with main_col1: st.metric("當月預計總獎金", f"{res[0]} 元")
        with main_col2:
            if revenue_tier != "不列入計算": st.info(f"包含個人業績獎金 ({revenue_tier}): {res[8]} 元")
        
        if st.button("產生並下載結算報表"):
            if not name or name.strip() == "": st.error("請輸入小編姓名")
            else:
                meta = {"館別": gym, "小編姓名": name, "報表日期": date_str}
                excel_file = generate_matrix_excel(
                    meta, res[1], res[0], deal_dict, extra_cls, loyalty_dict, upgrade_dict,
                    res[4], res[3], res[5], res[2], res[6], res[7], 
                    "正職" if is_ft else "兼職", brand_input, res[8], revenue_tier,
                    si_to_st_input, res[9]
                )
                st.download_button(
                    label="點我儲存 Excel 檔案",
                    data=excel_file,
                    file_name=f"{name}_獎金結算_{date_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # ===== 教練薪資結算系統 =====
    elif st.session_state.feature == "coach_salary":
        def generate_perfect_salary_report(uploaded_files, special_bonus_df):
            COACH_PRICING = {
                "林意潔": {"團1-2人": 800, "團3人": 900, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
                "陳秀蓉": {"團1-2人": 800, "團3人": 850, "團4人": 850, "團5人": 900, "團6人": 950, "1對2(1.5hr)": 1200, "1對2": 800, "1對1(1.5hr)": 1200, "1對1": 800},
                "陳怡廷": {"團1-2人": 800, "團3人": 850, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
                "鍾佳蓁 Rita": {"團1-2人": 900, "團3人": 950, "團4人": 1000, "團5人": 1050, "團6人": 1100, "1對2(1.5hr)": 1425, "1對2": 950, "1對1(1.5hr)": 1425, "1對1": 950},
                "黃宛婷": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "楊子慧(小在)": {"團1-2人": 650, "團3人": 800, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1050, "1對2": 700, "1對1(1.5hr)": 1050, "1對1": 700},
                "許力尹 LOUIS": {"團1-2人": 600, "團3人": 800, "團4人": 800, "團5人": 800, "團6人": 800, "1對2(1.5hr)": 1200, "1對2": 800, "1對1(1.5hr)": 1200, "1對1": 800},
                "顥顥": {"團1-2人": 0, "團3人": 1000, "團4人": 1000, "團5人": 1100, "團6人": 1100, "1對2(1.5hr)": 1500, "1對2": 1000, "1對1(1.5hr)": 1500, "1對1": 1000},
                "洪睿絃": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "紀儒蓁": {"團1-2人": 700, "團3人": 800, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "李翎瑋": {"團1-2人": 800, "團3人": 850, "團4人": 900, "團5人": 950, "團6人": 1000, "1對2(1.5hr)": 1275, "1對2": 850, "1對1(1.5hr)": 1275, "1對1": 850},
                "郭奕伶": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "郭品均": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "邴妍語": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "張鈞弼": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 850, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "蕭竣升": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "紀萃文": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "李函豫": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650},
                "尤子綺": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "張楷翌": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "侯懿庭": {"團1-2人": 700, "團3人": 750, "團4人": 800, "團5人": 850, "團6人": 900, "1對2(1.5hr)": 1125, "1對2": 750, "1對1(1.5hr)": 1125, "1對1": 750},
                "謝俐池": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "黃姿菁": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "籃郁雯": {"團1-2人": 550, "團3人": 600, "團4人": 700, "團5人": 700, "團6人": 700, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "徐漫": {"團1-2人": 500, "團3人": 550, "團4人": 600, "團5人": 650, "團6人": 700, "1對2(1.5hr)": 825, "1對2": 550, "1對1(1.5hr)": 825, "1對1": 550},
                "鄭筠馨": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "高舒涵": {"團1-2人": 550, "團3人": 600, "團4人": 650, "團5人": 700, "團6人": 750, "1對2(1.5hr)": 900, "1對2": 600, "1對1(1.5hr)": 900, "1對1": 600},
                "邱靜瑜": {"團1-2人": 600, "團3人": 650, "團4人": 700, "團5人": 750, "團6人": 800, "1對2(1.5hr)": 975, "1對2": 650, "1對1(1.5hr)": 975, "1對1": 650}
            }
            NAME_CONVERSION = {
                "意潔": "林意潔", "Cammy": "陳怡廷", "Vivi": "陳秀蓉", "怡廷": "陳怡廷", "佳蓁": "鍾佳蓁 Rita", "宛婷": "黃宛婷", "WanTing": "黃宛婷",
                "小在": "楊子慧(小在)", "Jae": "楊子慧(小在)", "LOUIS": "許力尹 LOUIS", "顥顥": "顥顥", "睿絃": "洪睿絃", "儒蓁": "紀儒蓁",
                "翎瑋": "李翎瑋", "奕伶": "郭奕伶", "品均": "郭品均", "妍語": "邴妍語", "鈞弼": "張鈞弼",
                "竣升": "蕭竣升", "萃萃": "紀萃文", "函豫": "李函豫", "Hanny": "李函豫", "子綺": "尤子綺", "Yuli": "尤子綺", "楷翌": "張楷翌", "Eric": "張楷翌",
                "懿庭": "侯懿庭", "Yvonne Hou": "侯懿庭", "俐池": "謝俐池", "Grace Hsieh": "謝俐池", "姿菁": "黃姿菁", "郁雯": "籃郁雯", "徐漫": "徐漫", "mandy": "徐漫",
                "筠馨": "鄭筠馨", "舒涵": "高舒涵", "靜瑜": "邱靜瑜"
            }
            SENIORITY_DATA = {
                "林意潔": "兩年以上", "陳秀蓉": "兩年以上", "陳怡廷": "兩年以上", "鍾佳蓁 Rita": "兩年以上",
                "黃宛婷": "一年以上至兩年以下", "楊子慧(小在)": "兩年以上", "許力尹 LOUIS": "一年以上至兩年以下",
                "顥顥": "一年以上至兩年以下", "洪睿絃": "一年以上至兩年以下", "紀儒蓁": "兩年以上",
                "李翎瑋": "一年以上至兩年以下", "郭奕伶": "一年以上至兩年以下", "郭品均": "一年以上至兩年以下",
                "邴妍語": "一年以上至兩年以下", "張鈞弼": "一年以上至兩年以下", "蕭竣升": "一年以上至兩年以下",
                "紀萃文": "一年以上至兩年以下", "李函豫": "一年以上至兩年以下", "尤子綺": "一年以下",
                "張楷翌": "一年以上至兩年以下", "侯懿庭": "一年以下", "謝俐池": "一年以下",
                "黃姿菁": "一年以下", "籃郁雯": "一年以下", "徐漫": "一年以下",
                "鄭筠馨": "一年以下", "高舒涵": "一年以下", "邱靜瑜": "一年以下"
            }
            
            display_course_types = ["團1-2人", "團3人", "團4人", "團5人", "團6人", "1對2(1.5hr)", "1對2", "1對1(1.5hr)", "1對1"]
            
            columns = [
                "教練_L", "課程_L", "單價_L", "義昌館堂數", "義昌館金額", "高美館堂數", "高美館金額", "中山館堂數", "中山館金額", "三館總堂數", "三館總金額",
                "應付金額_L", "三館獎金(春不老)_堂數達標獎金_L", "三館特別獎金_L", "總計_L", "執行業務_L", "補充保費_L", "應付薪資_L",
                "教練_R", "課程_R", "單價_R", "巨蛋館堂數", "巨蛋館金額",
                "應付金額_R", "巨蛋獎金_堂數達標獎金_R", "巨蛋特別獎金_R", "總計_R", "執行業務_R", "補充保費_R", "應付薪資_R", "備註"
            ]
            
            all_rows = []
            for coach in COACH_PRICING.keys():
                for course in display_course_types:
                    row = {col: 0 for col in columns}
                    row["教練_L"], row["課程_L"], row["單價_L"] = coach, course, COACH_PRICING[coach].get(course, 0)
                    row["教練_R"], row["課程_R"], row["單價_R"] = coach, course, COACH_PRICING[coach].get(course, 0)
                    row["備註"] = ""
                    all_rows.append(row)
            df_master = pd.DataFrame(all_rows)
            
            for file in uploaded_files:
                file.seek(0)
                df_info = pd.read_excel(file, sheet_name='統計總表', nrows=1, header=None)
                loc_name = str(df_info.iloc[0, 1]).strip()
                
                file.seek(0)
                df_raw = pd.read_excel(file, sheet_name='統計總表', skiprows=3).fillna(0)
                df_raw = df_raw.set_index(df_raw.columns[0]).transpose()
                df_raw.index.name = '姓名'
                df_stats = df_raw.reset_index()

                for _, s_row in df_stats.iterrows():
                    raw_excel_name = str(s_row['姓名']).strip()
                    if raw_excel_name == '合計' or not raw_excel_name or 'Unnamed' in raw_excel_name:
                        continue
                    mapped_name = None
                    if raw_excel_name in COACH_PRICING:
                        mapped_name = raw_excel_name
                    else:
                        mapped_name = NAME_CONVERSION.get(raw_excel_name)
                        if not mapped_name:
                            for coach_full_name in COACH_PRICING.keys():
                                if raw_excel_name in coach_full_name or coach_full_name in raw_excel_name:
                                    mapped_name = coach_full_name
                                    break
                    if not mapped_name: continue
                    coach_mask = (df_master["教練_L"] == mapped_name)
                    for course in display_course_types:
                        val = 0
                        if course == "團1-2人":
                            val = s_row.get("團1人", 0) + s_row.get("團2人", 0)
                        else:
                            val = s_row.get(course, 0)
                        if val == 0: continue
                        target_mask = coach_mask & (df_master["課程_L"] == course)
                        unit_price = COACH_PRICING.get(mapped_name, {}).get(course, 0)
                        if loc_name == "巨蛋館":
                            df_master.loc[target_mask, "巨蛋館堂數"] = val
                            df_master.loc[target_mask, "巨蛋館金額"] = val * unit_price
                        else:
                            s_col, a_col = f"{loc_name}堂數", f"{loc_name}金額"
                            if s_col in df_master.columns:
                                df_master.loc[target_mask, s_col] += val
                                df_master.loc[target_mask, a_col] += val * unit_price

            bonus_summary = []
            coach_bonus_map_s3 = {} 
            coach_bonus_map_sd = {} 
            cat1_list = ["團1-2人", "1對2(1.5hr)", "1對2", "1對1(1.5hr)", "1對1"]
            cat2_list = ["團3人", "團4人", "團5人", "團6人"]
            RATIO_CONFIG = {
                "兩年以上": {"門檻": 110, "1-2人": 0.64, "3-6人": 0.36},
                "一年以上至兩年以下": {"門檻": 90, "1-2人": 0.67, "3-6人": 0.33},
                "一年以下": {"門檻": 50, "1-2人": 0.60, "3-6人": 0.40}
            }
            for coach in COACH_PRICING.keys():
                c_data = df_master[df_master["教練_L"] == coach]
                seniority = SENIORITY_DATA.get(coach, "一年以下")
                config = RATIO_CONFIG.get(seniority, RATIO_CONFIG["一年以下"])
                total_classes = c_data[["義昌館堂數", "高美館堂數", "中山館堂數", "巨蛋館堂數"]].sum().sum()
                actual_sum1 = c_data.loc[c_data["課程_L"].isin(cat1_list), ["義昌館堂數", "高美館堂數", "中山館堂數", "巨蛋館堂數"]].sum().sum()
                actual_sum2 = c_data.loc[c_data["課程_L"].isin(cat2_list), ["義昌館堂數", "高美館堂數", "中山館堂數", "巨蛋館堂數"]].sum().sum()
                actual_ratio1 = actual_sum1 / total_classes if total_classes > 0 else 0
                actual_ratio2 = actual_sum2 / total_classes if total_classes > 0 else 0
                detail_text, b1, b2, b1_s3, b1_sd, b2_s3, b2_sd, ex_classes_cat1, ex_classes_cat2 = "", 0, 0, 0, 0, 0, 0, 0, 0
                if total_classes > config["門檻"]:
                    excess_total = total_classes - config["門檻"]
                    if actual_ratio1 >= config["1-2人"]:
                        ex_classes_cat1 = round(excess_total * actual_ratio1)
                        b1 = (min(ex_classes_cat1, 10) * 40) + (max(0, ex_classes_cat1 - 10) * 50)
                        subset1 = c_data[c_data["課程_L"].isin(cat1_list)]
                        s1_s3, s1_sd = subset1[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum().sum(), subset1["巨蛋館堂數"].sum()
                        s1_total = s1_s3 + s1_sd
                        if s1_total > 0:
                            b1_s3, b1_sd = round(b1 * (s1_s3 / s1_total)), round(b1 * (s1_sd / s1_total))
                    else: detail_text += "1-2人佔比未達標; "
                    if actual_ratio2 >= config["3-6人"]:
                        ex_classes_cat2 = round(excess_total * actual_ratio2)
                        if seniority == "兩年以上": b2 = (min(ex_classes_cat2, 10) * 60) + (max(0, ex_classes_cat2 - 10) * 80)
                        else: b2 = (min(ex_classes_cat2, 10) * 50) + (min(max(0, ex_classes_cat2 - 10), 10) * 80) + (max(0, ex_classes_cat2 - 20) * 100)
                        subset2 = c_data[c_data["課程_L"].isin(cat2_list)]
                        s2_s3, s2_sd = subset2[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum().sum(), subset2["巨蛋館堂數"].sum()
                        s2_total = s2_s3 + s2_sd
                        if s2_total > 0:
                            b2_s3, b2_sd = round(b2 * (s2_s3 / s2_total)), round(b2 * (s2_sd / s2_total))
                    else: detail_text += "3-6人佔比未達標; "
                    if seniority == "一年以下":
                        extra_bonus = 2000 if excess_total >= 5 else 1000
                        if (b1 + b2) < extra_bonus:
                            detail_text += f"取保底 {extra_bonus}; "
                            b1, b2 = extra_bonus * config["1-2人"], extra_bonus * config["3-6人"]
                            subset1, subset2 = c_data[c_data["課程_L"].isin(cat1_list)], c_data[c_data["課程_L"].isin(cat2_list)]
                            s1_s3, s1_sd = subset1[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum().sum(), subset1["巨蛋館堂數"].sum()
                            s2_s3, s2_sd = subset2[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum().sum(), subset2["巨蛋館堂數"].sum()
                            if (s1_s3 + s1_sd) > 0: b1_s3, b1_sd = round(b1 * (s1_s3 / (s1_s3 + s1_sd))), round(b1 * (s1_sd / (s1_s3 + s1_sd)))
                            if (s2_s3 + s2_sd) > 0: b2_s3, b2_sd = round(b2 * (s2_s3 / (s2_s3 + s2_sd))), round(b2 * (s2_sd / (s2_s3 + s2_sd)))
                coach_bonus_map_s3[coach], coach_bonus_map_sd[coach] = round(b1_s3 + b2_s3), round(b1_sd + b2_sd)
                bonus_val = round(b1 + b2)
                if not detail_text: detail_text = f"總超額 {round(excess_total if total_classes > config['門檻'] else 0)} 堂"
                for label, categories in [("1人至2人", cat1_list), ("3人至6人", cat2_list)]:
                    subset = c_data[c_data["課程_L"].isin(categories)]
                    s3, sd = round(subset[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum().sum()), round(subset["巨蛋館堂數"].sum())
                    st_val = s3 + sd
                    bonus_summary.append({
                        "教練姓名": coach, "課程人數": label,"三館總堂數": s3, "巨蛋館堂數": sd, "四館總堂數": st_val,
                        "級距達標門檻": f"{round(config['1-2人']*100 if label=='1人至2人' else config['3-6人']*100)}%", 
                        "個別超標堂數": round(ex_classes_cat1 if label == "1人至2人" else ex_classes_cat2),
                        "總堂數門檻": config["門檻"] if label == "1人至2人" else "", "計算明細": detail_text if label == "1人至2人" else "",
                        "級距分配獎金": round(b1 if label == "1人至2人" else b2), "合計總獎金": bonus_val if label == "1人至2人" else "",
                        "三館獎金(春不老)": b1_s3 if label == "1人至2人" else b2_s3, "巨蛋獎金": b1_sd if label == "1人至2人" else b2_sd, "年資": seniority
                    })
                bonus_summary.append({k: "" for k in ["教練姓名", "四館總堂數", "個別超標堂數", "總堂數門檻", "計算明細", "級距分配獎金", "合計總獎金", "三館獎金(春不老)", "巨蛋獎金", "年資"]})
            
            final_dfs, spec_bonus_map = [], special_bonus_df.set_index("教練姓名").to_dict('index')
            for coach in COACH_PRICING.keys():
                c_df = df_master[df_master["教練_L"] == coach].copy()
                bonus_s3, bonus_sd = coach_bonus_map_s3.get(coach, 0), coach_bonus_map_sd.get(coach, 0)
                s_bonus_s3, s_bonus_sd = int(spec_bonus_map.get(coach, {}).get("三館特別獎金(加項)", 0)), int(spec_bonus_map.get(coach, {}).get("巨蛋特別獎金(加項)", 0))
                remark = spec_bonus_map.get(coach, {}).get("備註", "")
                c_df["三館總堂數"], c_df["三館總金額"] = c_df[["義昌館堂數", "高美館堂數", "中山館堂數"]].sum(axis=1), c_df[["義昌館金額", "高美館金額", "中山館金額"]].sum(axis=1)
                total_L_lessons = round(c_df["三館總金額"].sum())
                total_L_sum = total_L_lessons + bonus_s3 + s_bonus_s3
                tax_L, health_L = (round(total_L_sum * 0.1) if total_L_sum >= 20000 else 0), (round(total_L_sum * 0.0211) if total_L_sum >= 20000 else 0)
                pay_L = total_L_sum - tax_L - health_L
                total_R_lessons = round(c_df["巨蛋館金額"].sum())
                total_R_sum = total_R_lessons + bonus_sd + s_bonus_sd
                tax_R, health_R = (round(total_R_sum * 0.1) if total_R_sum >= 20000 else 0), (round(total_R_sum * 0.0211) if total_R_sum >= 20000 else 0)
                pay_R = total_R_sum - tax_R - health_R
                c_df["應付金額_L"], c_df["三館獎金(春不老)_堂數達標獎金_L"], c_df["三館特別獎金_L"], c_df["總計_L"] = total_L_lessons, bonus_s3, s_bonus_s3, total_L_sum
                c_df["執行業務_L"], c_df["補充保費_L"], c_df["應付薪資_L"] = tax_L, health_L, pay_L
                c_df["應付金額_R"], c_df["巨蛋獎金_堂數達標獎金_R"], c_df["巨蛋特別獎金_R"], c_df["總計_R"] = total_R_lessons, bonus_sd, s_bonus_sd, total_R_sum
                c_df["執行業務_R"], c_df["補充保費_R"], c_df["應付薪資_R"], c_df["備註"] = tax_R, health_R, pay_R, remark
                final_dfs.append(c_df)
                sub_row = {col: 0 for col in columns}
                sub_row["教練_L"], sub_row["課程_L"], sub_row["教練_R"], sub_row["課程_R"] = coach, "小計", coach, "小計"
                for gym in ["義昌館", "高美館", "中山館", "巨蛋館"]:
                    sub_row[f"{gym}堂數"], sub_row[f"{gym}金額"] = round(c_df[f"{gym}堂數"].sum()), round(c_df[f"{gym}金額"].sum())
                sub_row["三館總堂數"], sub_row["三館總金額"] = round(c_df["三館總堂數"].sum()), round(c_df["三館總金額"].sum())
                sub_row["應付金額_L"], sub_row["三館獎金(春不老)_堂數達標獎金_L"], sub_row["三館特別獎金_L"], sub_row["總計_L"] = total_L_lessons, bonus_s3, s_bonus_s3, total_L_sum
                sub_row["執行業務_L"], sub_row["補充保費_L"], sub_row["應付薪資_L"] = tax_L, health_L, pay_L
                sub_row["應付金額_R"], sub_row["巨蛋獎金_堂數達標獎金_R"], sub_row["巨蛋特別獎金_R"], sub_row["總計_R"] = total_R_lessons, bonus_sd, s_bonus_sd, total_R_sum
                sub_row["執行業務_R"], sub_row["補充保費_R"], sub_row["應付薪資_R"], sub_row["備註"] = tax_R, health_R, pay_R, remark
                final_dfs.append(pd.DataFrame([sub_row]))
            df_final = pd.concat(final_dfs, ignore_index=True)[columns]
            subtotal_rows = df_final[df_final["課程_L"] == "小計"]
            three_bonus_pool, dome_bonus_pool = subtotal_rows["三館獎金(春不老)_堂數達標獎金_L"].sum(), subtotal_rows["巨蛋獎金_堂數達標獎金_R"].sum()
            yi_amt, gao_amt, zhong_amt, dome_amt = subtotal_rows["義昌館金額"].sum(), subtotal_rows["高美館金額"].sum(), subtotal_rows["中山館金額"].sum(), subtotal_rows["巨蛋館金額"].sum()
            three_amt_sum, four_amt_sum = yi_amt + gao_amt + zhong_amt, yi_amt + gao_amt + zhong_amt + dome_amt
            yi_cnt, gao_cnt, zhong_cnt, dome_cnt = subtotal_rows["義昌館堂數"].sum(), subtotal_rows["高美館堂數"].sum(), subtotal_rows["中山館堂數"].sum(), subtotal_rows["巨蛋館堂數"].sum()
            three_cnt_sum, four_cnt_sum = yi_cnt + gao_cnt + zhong_cnt, yi_cnt + gao_cnt + zhong_cnt + dome_cnt
            def calc_ratio_str(part, total): return f"{ (part / total * 100):.2f}%" if total > 0 else "0.00%"
            def get_bonus_share(part_cnt, total_cnt, total_bonus): return round(total_bonus * (part_cnt / total_cnt)) if total_cnt > 0 else 0
            summary_rows = [
                ["義昌:", yi_amt, round(yi_amt * 1.05), calc_ratio_str(yi_amt, four_amt_sum), yi_cnt, calc_ratio_str(yi_cnt, four_cnt_sum), get_bonus_share(yi_cnt, three_cnt_sum, three_bonus_pool), "", 0, 0],
                ["高美:", gao_amt, round(gao_amt * 1.05), calc_ratio_str(gao_amt, four_amt_sum), gao_cnt, calc_ratio_str(gao_cnt, four_cnt_sum), get_bonus_share(gao_cnt, three_cnt_sum, three_bonus_pool), "", 0, 0],
                ["中山:", zhong_amt, round(zhong_amt * 1.05), calc_ratio_str(zhong_amt, four_amt_sum), zhong_cnt, calc_ratio_str(zhong_cnt, four_cnt_sum), get_bonus_share(zhong_cnt, three_cnt_sum, three_bonus_pool), "", 0, 0],
                ["三館合計:", three_amt_sum, round(three_amt_sum * 1.05), calc_ratio_str(three_amt_sum, four_amt_sum), three_cnt_sum, calc_ratio_str(three_cnt_sum, four_cnt_sum), three_bonus_pool, "", 0, 0],
                ["巨蛋:", dome_amt, round(dome_amt * 1.05), calc_ratio_str(dome_amt, four_amt_sum), dome_cnt, calc_ratio_str(dome_cnt, four_cnt_sum), dome_bonus_pool, "", 0, 0],
                ["四館合計:", four_amt_sum, round(four_amt_sum * 1.05), calc_ratio_str(four_amt_sum, four_amt_sum), four_cnt_sum, calc_ratio_str(four_cnt_sum, four_cnt_sum), three_bonus_pool + dome_bonus_pool, "", 0, 0]
            ]
            df_summary = pd.DataFrame(summary_rows, columns=["", "營收(未稅)", "營收(含稅)", "營收比例", "堂數", "堂數比例", "堂數達標獎金", "各館薪資比例", "教練平均單價", "完課人數"])
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='薪資清單')
                pd.DataFrame(bonus_summary).to_excel(writer, index=False, sheet_name='堂數達標獎金')
                df_summary.to_excel(writer, index=False, sheet_name='統計總表')
                thin, center = Side(style='thin'), Alignment(horizontal='center', vertical='center')
                border, header_fill, sub_fill, red_font = Border(top=thin, left=thin, right=thin, bottom=thin), PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"), Font(color="FF0000", bold=True)
                ws = writer.sheets['薪資清單']
                display_headers = ["教練", "課程", "單價", "義昌館堂數", "義昌館金額", "高美館堂數", "高美館金額", "中山館堂數", "中山館金額", "三館總堂數", "三館總金額", "應付金額", "三館獎金(春不老)_堂數達標獎金(加項)", "三館特別獎金(加項)", "總計", "執行業務(扣款)", "補充保費(扣款)", "應付薪資", "教練", "課程", "單價", "巨蛋館堂數", "巨蛋館金額", "應付金額", "巨蛋獎金_堂數達標獎金(加項)", "巨蛋特別獎金(加項)", "總計", "執行業務(扣款)", "補充保費(扣款)", "應付薪資", "備註"]
                for i, h in enumerate(display_headers): ws.cell(row=1, column=i+1).value = h
                for cell in ws[1]: cell.fill, cell.font, cell.alignment, cell.border = header_fill, Font(bold=True), center, border
                merge_cols, start_r = [1, 12, 13, 14, 15, 16, 17, 18, 19, 24, 25, 26, 27, 28, 29, 30, 31], 2
                for r in range(2, ws.max_row + 1):
                    curr_coach, curr_course = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
                    next_coach = ws.cell(row=r + 1, column=1).value if r < ws.max_row else None
                    is_sub = (curr_course == "小計")
                    for c in range(1, ws.max_column + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.border, cell.alignment = border, center
                        if is_sub: cell.fill = sub_fill
                    if next_coach != curr_coach:
                        if r >= start_r:
                            for c in merge_cols:
                                ws.merge_cells(start_row=start_r, start_column=c, end_row=r, end_column=c)
                                ws.cell(row=start_r, column=c).alignment = center
                        start_r = r + 1
                for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 12
                ws.column_dimensions['M'].width, ws.column_dimensions['X'].width, ws.column_dimensions['AE'].width = 35, 30, 30
                ws_bonus = writer.sheets['堂數達標獎金']
                for row in ws_bonus.iter_rows(min_row=1, max_row=ws_bonus.max_row):
                    for cell in row:
                        cell.alignment, cell.border = center, border
                        if cell.row == 1: cell.fill, cell.font = header_fill, Font(bold=True)
                for col in ws_bonus.columns: ws_bonus.column_dimensions[col[0].column_letter].width = 15
                ws_bonus.column_dimensions['J'].width = 50
                ws_s = writer.sheets['統計總表']
                for r_idx, row in enumerate(ws_s.iter_rows(min_row=1, max_row=ws_s.max_row), 1):
                    for cell in row:
                        cell.border, cell.alignment = border, center
                        if r_idx == 1: cell.fill, cell.font = header_fill, Font(bold=True)
                        if "合計" in str(ws_s.cell(row=r_idx, column=1).value): cell.font = red_font
                for col in ws_s.columns: ws_s.column_dimensions[col[0].column_letter].width = 16
            return output.getvalue()

        st.title("教練薪資統計")
        COACH_NAMES = ["林意潔", "陳秀蓉", "陳怡廷", "鍾佳蓁 Rita", "黃宛婷", "楊子慧(小在)", "許力尹 LOUIS", "顥顥", "洪睿絃", "紀儒蓁", "李翎瑋", "郭奕伶", "郭品均", "邴妍語", "張鈞弼", "蕭竣升", "紀萃文", "李函豫", "尤子綺", "張楷翌", "侯懿庭", "謝俐池", "黃姿菁", "籃郁雯", "徐漫", "鄭筠馨", "高舒涵", "邱靜瑜"]
        st.subheader("1. 上傳資料")
        uploaded_files = st.file_uploader("選擇原始 Excel 檔案", type=["xlsx"], accept_multiple_files=True, key="coach_salary_uploader")
        st.subheader("2. 輸入特別獎金 (選填)")
        st.info("可在下方表格直接輸入各教練的特別獎金與備註（如說明會、觀課等)")
        if 'special_bonus_data' not in st.session_state:
            st.session_state.special_bonus_data = pd.DataFrame({"教練姓名": COACH_NAMES, "三館特別獎金(加項)": [0] * len(COACH_NAMES), "巨蛋特別獎金(加項)": [0] * len(COACH_NAMES), "備註": [""] * len(COACH_NAMES)})
        edited_df = st.data_editor(st.session_state.special_bonus_data, column_config={"教練姓名": st.column_config.TextColumn("教練姓名", disabled=True), "三館特別獎金(加項)": st.column_config.NumberColumn("三館特別獎金(加項)", format="%d"), "巨蛋特別獎金(加項)": st.column_config.NumberColumn("巨蛋特別獎金(加項)", format="%d"), "備註": st.column_config.TextColumn("備註")}, hide_index=True, use_container_width=True, key="coach_salary_editor")
        st.subheader("3. 產生報表")
        if uploaded_files:
            if st.button("開始計算薪資並排版", key="start_calc_btn"):
                try:
                    final_excel_data = generate_perfect_salary_report(uploaded_files, edited_df)
                    st.download_button(label="下載薪資明細表", data=final_excel_data, file_name="教練薪資明細.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_salary_btn")
                except Exception as e:
                    st.error(f"計算過程中發生錯誤: {e}")
        else:
            st.warning("請先上傳預約統計表檔案。")

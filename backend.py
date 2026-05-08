from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import io
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np

app = Flask(__name__)
CORS(app)

# 老師名稱轉換字典
NAME_CONVERSION = {
    "意潔": "林意潔", "Cammy": "陳怡廷", "Vivi": "陳秀蓉", "怡廷": "陳怡廷", 
    "秀蓉": "陳秀蓉", "佳蓁": "鍾佳蓁 Rita", "宛婷": "黃宛婷", "WanTing": "黃宛婷",
    "小在": "楊子慧(小在)", "Jae": "楊子慧(小在)", "LOUIS": "許力尹 LOUIS", 
    "許力尹": "許力尹 LOUIS", "顥顥": "顥顥", "睿絃": "洪睿絃", "儒蓁": "紀儒蓁",
    "翎瑋": "李翎瑋", "奕伶": "郭奕伶", "品均": "郭品均", "妍語": "邴妍語", 
    "鈞弼": "張鈞弼", "竣升": "蕭竣升", "萃萃": "紀萃文", "萃文": "紀萃文", 
    "函豫": "李函豫", "Hanny": "遊函豫", "子綺": "尤子綺", "Yuli": "尤子綺", 
    "楷翌": "張楷翌", "Eric": "張楷翌", "懿庭": "侯懿庭", "Yvonne Hou": "侯懿庭", 
    "俐池": "謝俐池", "Grace Hsieh": "謝俐池", "姿菁": "黃姿菁", "郁雯": "籃郁雯", 
    "徐漫": "徐漫", "mandy": "徐漫", "漫漫": "徐漫", "筠馨": "鄭筠馨", 
    "舒涵": "高舒涵", "靜瑜": "邱靜瑜"
}

TEACHER_ORDER = list(dict.fromkeys(NAME_CONVERSION.values()))

def get_formal_name(raw_name):
    """轉換老師名稱為正式名稱"""
    name_str = str(raw_name).strip()
    for key, formal_name in NAME_CONVERSION.items():
        if key.lower() in name_str.lower():
            return formal_name
    return name_str

def teacher_sort_key(name):
    """老師排序權重"""
    if name in TEACHER_ORDER:
        return TEACHER_ORDER.index(name)
    return len(TEACHER_ORDER)

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康檢查"""
    return jsonify({'status': 'ok'})

@app.route('/api/process-lesson-report', methods=['POST'])
def process_lesson_report():
    """處理預約報表"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '未找到檔案'}), 400
        
        file = request.files['file']
        branch = request.form.get('branch', '全部')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        # 讀取檔案
        if file.filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file, header=None, nrows=20)
            file.seek(0)
            target_row = 0
            for i, row in df.iterrows():
                row_str = " ".join([str(x) for x in row.values])
                if '課程日期' in row_str or '授課老師' in row_str:
                    target_row = i
                    break
            df = pd.read_excel(file, skiprows=target_row)
        else:
            encodings = ['utf-8-sig', 'big5', 'cp950', 'gbk']
            for enc in encodings:
                try:
                    file.seek(0)
                    df = pd.read_csv(file, encoding=enc)
                    if '課程日期' not in "".join(df.columns.astype(str)):
                        file.seek(0)
                        df = pd.read_csv(file, encoding=enc, skiprows=1)
                    break
                except:
                    continue
        
        if df is None or df.empty:
            return jsonify({'success': False, 'error': '無法辨識檔案格式'}), 400
        
        df.columns = df.columns.astype(str).str.strip()
        
        # 尋找欄位
        def find_col(possible_names):
            for name in possible_names:
                for col in df.columns:
                    if name in col and "Unnamed" not in col:
                        return col
            return None
        
        date_col = find_col(['課程日期', '日期'])
        teacher_col = find_col(['授課老師', '老師'])
        course_col = find_col(['課程名稱', '課程'])
        count_col = find_col(['預約總人數', '預約人數', '人數'])
        duration_col = find_col(['課程時數', '分鐘'])
        branch_col = find_col(['館別', '分館'])
        
        if not date_col or not teacher_col:
            return jsonify({'success': False, 'error': '缺少必要欄位'}), 400
        
        # 資料清洗
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        df = df.dropna(subset=[date_col])
        
        if branch_col and branch != "全部":
            df = df[df[branch_col].astype(str).str.contains(branch)]
        
        start = pd.to_datetime(start_date)
        end = pd.to_datetime(end_date)
        df = df[(df[date_col].dt.date >= start.date()) & (df[date_col].dt.date <= end.date())]
        
        df[count_col] = pd.to_numeric(df[count_col], errors='coerce').fillna(0)
        df_filtered = df[df[count_col] > 0].copy()
        
        # 轉換正式姓名
        df_filtered['正式姓名'] = df_filtered[teacher_col].apply(get_formal_name)
        
        # 統計
        stats_items = [
            '團1人', '團2人', '團3人', '團4人', '團5人', '團6人',
            '1對2(1.5hr)', '1對2', '1對1(1.5hr)', '1對1', '觀課'
        ]
        
        all_formal_teachers = df_filtered['正式姓名'].unique().tolist()
        df_stats = pd.DataFrame(0, index=all_formal_teachers, columns=stats_items)
        
        for _, row in df_filtered.iterrows():
            teacher = row['正式姓名']
            course_name = str(row[course_col]).strip()
            count = int(row[count_col])
            duration = row[duration_col] if duration_col else 60
            
            if '觀課' in course_name:
                df_stats.at[teacher, '觀課'] += 1
            elif '一對一' in course_name:
                if duration >= 90: df_stats.at[teacher, '1對1(1.5hr)'] += 1
                else: df_stats.at[teacher, '1對1'] += 1
            elif '一對二' in course_name:
                if duration >= 90: df_stats.at[teacher, '1對2(1.5hr)'] += 1
                else: df_stats.at[teacher, '1對2'] += 1
            else:
                if 1 <= count <= 6:
                    col_name = f'團{count}人'
                    df_stats.at[teacher, col_name] += 1
        
        df_stats['小計'] = df_stats.sum(axis=1)
        
        # 排序
        df_stats['sort_key'] = df_stats.index.map(teacher_sort_key)
        df_stats = df_stats.sort_values('sort_key').drop(columns=['sort_key'])
        
        # 合計
        total_row = df_stats.sum().to_frame().T
        total_row.index = ['合計']
        df_final = pd.concat([df_stats, total_row])
        
        # 轉置
        df_transposed = df_final.T
        
        # 準備回應
        teachers = [col for col in df_transposed.columns if col != '合計'] + ['合計']
        stats_data = []
        for idx, row in df_transposed.iterrows():
            stats_data.append({
                'course': idx,
                'values': [int(row[t]) for t in teachers]
            })
        
        # 儲存到 session 以供下載
        app.lesson_report_data = {
            'df': df_transposed,
            'branch': branch,
            'start_date': start_date,
            'end_date': end_date
        }
        
        return jsonify({
            'success': True,
            'teachers': teachers,
            'stats': stats_data
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/process-sales-report', methods=['POST'])
def process_sales_report():
    """處理業績報表"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '未找到檔案'}), 400
        
        file = request.files['file']
        df_raw = pd.read_excel(file, header=1)
        
        # 建立基礎資料
        df_base = pd.DataFrame()
        df_base["合約建立日期日期"] = df_raw["交易日期"]
        df_base["業績人員"] = df_raw["銷售人員"]
        df_base["會員姓名"] = df_raw["會員姓名"]
        df_base["開課日期"] = np.nan
        df_base["合約類型"] = df_raw["票券/商品種類"]
        df_base["堂數"] = pd.to_numeric(df_raw["數量"], errors='coerce').fillna(0)
        
        total_price = pd.to_numeric(df_raw["總計(元)"], errors='coerce').fillna(0)
        df_base["合約總價"] = total_price
        df_base["金額(未稅)"] = (total_price / 1.05).round(0)
        
        # 計算單價
        def get_unit_price(contract_type):
            contract_type = str(contract_type)
            if '一對一' in contract_type: return 2300
            elif '一對二' in contract_type: return 1300
            elif '一對三' in contract_type: return 1050
            elif '一對四' in contract_type: return 900
            else: return 0
        
        df_base["購買合約原價"] = df_base["合約類型"].apply(get_unit_price) * df_base["堂數"]
        
        df_base["銷售折數"] = np.where(
            df_base["購買合約原價"] != 0,
            (df_base["金額(未稅)"] / df_base["購買合約原價"]).round(2),
            0
        )
        
        df_base["活動"] = df_raw["Tag"]
        
        cond = [
            (df_base["銷售折數"] > 0.8),
            (df_base["銷售折數"] == 0.8),
            (df_base["銷售折數"] < 0.8) & (df_base["銷售折數"] > 0)
        ]
        
        bonus_rate = np.select(cond, [0.02, 0.015, 0.01], default=0)
        
        df_base["業績計算"] = np.select(cond, ["2%", "1.5%", "1%"], default="0%")
        df_base["業績獎金"] = (df_base["金額(未稅)"] * bonus_rate).round(0)
        
        df_base["備註"] = df_raw["備註"].astype(str).replace('nan', '')
        
        target_columns = [
            "合約建立日期", "業績人員", "會員姓名", "開課日期", "合約類型", 
            "堂數", "合約總價", "金額(未稅)", "銷售折數", "活動", 
            "業績計算", "業績獎金", "購買合約原價", "備註"
        df_base = df_base[target_columns]
        # 分類
        df_new = df_base[df_base["備註"].str.contains("新購|首購")].copy()
        df_renew = df_base[df_base["備註"].str.contains("續購")].copy()
        # 排序
        df_new = df_new.sort_values(by="業績人員").reset_index(drop=True)
        df_renew = df_renew.sort_values(by="業績人員").reset_index(drop=True)
        # 儲存到 session
        app.sales_report_data = {
            'new': df_new,
            'renew': df_renew
        }
        
        # 準備預覽資料
        new_sales = df_new[['業績人員', '會員姓名', '金額(未稅)']].head(5).values.tolist()
        renew_sales = df_renew[['業績人員', '會員姓名', '金額(未稅)']].head(5).values.tolist()
        
        return jsonify({
            'success': True,
            'new_sales': [{'salesperson': row[0], 'member': row[1], 'amount': int(row[2])} for row in new_sales],
            'renew_sales': [{'salesperson': row[0], 'member': row[1], 'amount': int(row[2])} for row in renew_sales]
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download-bonus-report', methods=['POST'])
def download_bonus_report():
    """下載獎金結算報表"""
    try:
        data = request.json
        
        # 計算獎金
        deal_bonus = (data['deal_today'] * 80 + data['deal_48h'] * 60 + 
                     data['deal_7d'] * 50 + data['deal_over7'] * 0)
        class_bonus = data['extra_classes'] * 30
        loyalty_bonus = (data['loyalty_10'] * 100 + data['loyalty_20'] * 200 + 
                        data['loyalty_30'] * 300 + data['loyalty_40'] * 500)
        upgrade_bonus = (data['upgrade_1213'] * 100 + data['upgrade_group'] * 150 + 
                        data['upgrade_class'] * 300)
        
        # 品牌知名度
        base_val = 5 if data['employment_type'] == 'full-time' else 2
        if data['brand_count'] == 0:
            brand_bonus = -200
            brand_note = '推廣人數為 0'
        elif data['brand_count'] < base_val:
            brand_bonus = -100
            brand_note = f"未達門檻 ({base_val}位)"
        elif data['brand_count'] == base_val:
            brand_bonus = 0
            brand_note = '符合基本門檻'
        else:
            extra_units = (data['brand_count'] - base_val) // 5
            brand_bonus = extra_units * 200
            brand_note = f"加發 {extra_units} 組獎金"
        
        # 月高手獎勵
        total_deals = (data['deal_today'] + data['deal_48h'] + data['deal_7d'] + 
                      data['deal_over7'] + data['upgrade_1213'] + data['upgrade_group'] + 
                      data['upgrade_class'] + data['loyalty_10'] + data['loyalty_20'] + 
                      data['loyalty_30'] + data['loyalty_40'] + data['extra_classes'])
        
        if total_deals >= 50:
            monthly_bonus = 5000
        elif total_deals >= 30:
            monthly_bonus = 2000
        else:
            monthly_bonus = 0
        
        # 個人業績獎金
        revenue_bonus_map = {
            'none': 0,
            '120000': 2000,
            '240000': 4000,
            '300000': 6000
        }
        revenue_bonus = revenue_bonus_map.get(data['revenue_tier'], 0)
        
        total_bonus = (deal_bonus + class_bonus + loyalty_bonus + upgrade_bonus + 
                      brand_bonus + monthly_bonus + revenue_bonus)
        
        # 建立 Excel
        wb = Workbook()
        ws = wb.active
        ws.title = '獎金結算'
        
        bold_font = Font(bold=True, name='微軟正黑體')
        center_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 基本資訊
        basic_info = [
            ['館別', data['gym'], '', '小編姓名', data['name']],
            ['報表日期', data['date'], '', '員工身份', '正職' if data['employment_type'] == 'full-time' else '兼職'],
            ['個人業績級別', data['revenue_tier'] if data['revenue_tier'] != 'none' else '不列入計算', '', '', '']
        ]
        
        for r_idx, row in enumerate(basic_info, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if value:
                    cell.alignment = center_align
                    if c_idx in [1, 4]:
                        cell.font = bold_font
                        cell.fill = header_fill
                    cell.border = thin_border
        
        # 獎金矩陣
        ws.cell(row=5, column=1, value='項目').font = bold_font
        ws.cell(row=5, column=2, value='數據/次數').font = bold_font
        ws.cell(row=5, column=3, value='獎金金額').font = bold_font
        ws.cell(row=5, column=4, value='備註').font = bold_font
        
        matrix_data = [
            ['個人業績獎金', data['revenue_tier'] if data['revenue_tier'] != 'none' else '不列入', revenue_bonus, ''],
            ['體驗成交', data['deal_today'] + data['deal_48h'] + data['deal_7d'] + data['deal_over7'], deal_bonus, ''],
            ['補位獎金', data['extra_classes'], class_bonus, ''],
            ['回流獎金', data['loyalty_10'] + data['loyalty_20'] + data['loyalty_30'] + data['loyalty_40'], loyalty_bonus, ''],
            ['結構升級獎金', data['upgrade_1213'] + data['upgrade_group'] + data['upgrade_class'], upgrade_bonus, ''],
            ['品牌知名度獎金', data['brand_count'], brand_bonus, brand_note],
            ['月高手獎勵', total_deals, monthly_bonus, f'總轉換筆數: {total_deals}'],
            ['總計', '', total_bonus, '']
        ]
        
        for r_idx, row_data in enumerate(matrix_data, 6):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border
                if r_idx == 6 + len(matrix_data) - 1:
                    cell.font = bold_font
        
        # 調整欄寬
        for i in range(1, 5):
            ws.column_dimensions[get_column_letter(i)].width = 20
        
        # 保存到記憶體
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{data['name']}_獎金結算_{data['date']}.xlsx"
        )
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
